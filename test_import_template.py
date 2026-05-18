#!/usr/bin/env python3
"""
Test script pour générer un fichier Excel de test pour l'import.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_test_excel():
    """Crée un fichier Excel de test avec des données valides."""
    
    # Créer un workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stagiaires"
    
    # En-têtes
    headers = ["prenom", "nom", "departement", "type_stagiaire", "ecole", "annee_etudes", "date_debut", "date_fin", "email"]
    ws.append(headers)
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Données de test
    test_data = [
        ["Ahmed", "Hassan", "Cardiologie", "Médecin", "Université de Marrakech", "4ème année", "2026-01-15", "2026-03-15", "ahmed.hassan@email.com"],
        ["Fatima", "Bennani", "Pédiatrie", "Infirmier", "École d'Infirmiers", "2ème année", "2026-02-01", "2026-04-01", "fatima.bennani@email.com"],
        ["Mohamed", "Amara", "Cardiologie", "Médecin", "Université Casablanca", "3ème année", "2026-01-20", "2026-03-20", "mohamed.amara@email.com"],
        ["Leila", "Chraibi", "Pédiatrie", "Infirmier", "Lycée technique", "1ère année", "2026-02-15", "2026-05-15", "leila.chraibi@email.com"],
        ["Tarek", "Bennani", "Chirurgie", "Administratif", "Institut administratif", "Diplômé", "2026-03-01", "2026-05-01", "tarek.bennani@email.com"],
    ]
    
    # Ajouter les données
    for row in test_data:
        ws.append(row)
    
    # Largeurs de colonne
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 20
    
    # Sauvegarder
    filename = "c:\\CHU_PFE_clean\\test_import_data.xlsx"
    wb.save(filename)
    print(f"✅ Fichier Excel de test créé : {filename}")
    return filename

if __name__ == "__main__":
    create_test_excel()

"""
excel_export.py
===============
Génère automatiquement un fichier Excel structuré depuis la base de données.

Utilisation :
  - Appelé automatiquement après chaque scan via main.py
  - Appelé manuellement : python excel_export.py
  - Appelé depuis le dashboard (bouton télécharger)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from database import SessionLocal, Intern, DailyStatus, Department
import os

BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BACKEND_DIR)
EXPORT_PATH = os.path.join(PROJECT_DIR, "Datas", "CHU_Pointages.xlsx")

# ── COULEURS ──────────────────────────────────────────────────────────────────
BLUE_HEADER  = "1E88E5"
GREEN        = "C8E6C9"
RED          = "FFCDD2"
ORANGE       = "FFE0B2"
GREY_ROW     = "F5F7FA"
WHITE        = "FFFFFF"

STATUS_COLORS = {
    "on_time":        "C8E6C9",   # vert
    "late":           "FFE0B2",   # orange
    "missed_checkin": "FFCDD2",   # rouge clair
    "early_checkout": "FFE0B2",   # orange
    "missed_checkout":"FFCDD2",   # rouge clair
    "absent":         "EF9A9A",   # rouge
    "Présent":        "C8E6C9",
    "Retard":         "FFE0B2",
}

def make_border():
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row_num, n_cols, bg=BLUE_HEADER):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill    = PatternFill("solid", fgColor=bg)
        cell.font    = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border  = make_border()

def style_data_row(ws, row_num, n_cols, bg=WHITE):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = make_border()

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


# ── DATA FETCH ────────────────────────────────────────────────────────────────

def fetch_data(department_id: str = None):
    """
    Fetch all data, or filter by department_id for supervisors.
    department_id=None  → super_admin, gets everything
    department_id=UUID  → supervisor, gets only their department
    """
    db = SessionLocal()
    try:
        query = (
            db.query(
                Intern.id,
                Intern.first_name,
                Intern.last_name,
                Department.name.label("department"),
            )
            .join(Department, Intern.department_id == Department.id)
        )
        # Filter by department if supervisor
        if department_id:
            query = query.filter(Intern.department_id == department_id)

        rows = query.all()
        df_interns = pd.DataFrame(rows, columns=["id","first_name","last_name","department"])
        df_interns["name"] = df_interns["first_name"] + " " + df_interns["last_name"]

        # Daily status — filtered to the interns we fetched
        intern_ids = df_interns["id"].tolist()
        daily_query = db.query(DailyStatus)
        if department_id:
            daily_query = daily_query.filter(DailyStatus.intern_id.in_(intern_ids))

        daily = daily_query.all()
        df_daily = pd.DataFrame([{
            "intern_id":       d.intern_id,
            "date":            d.date,
            "arrival_time":    d.arrival_time,
            "departure_time":  d.departure_time,
            "status":          d.status,
            "checkin_status":  d.checkin_status,
            "checkout_status": d.checkout_status,
            "work_duration":   d.work_duration,
            "needs_attention": d.needs_attention,
        } for d in daily])

        return df_interns, df_daily
    finally:
        db.close()


# ── SHEET 1 — POINTAGES JOURNALIERS ──────────────────────────────────────────

def sheet_pointages(wb, df_interns, df_daily):
    ws = wb.create_sheet("Pointages Journaliers")
    ws.sheet_view.showGridLines = False

    if df_daily.empty:
        ws["A1"] = "Aucune donnée de pointage disponible."
        return

    merged = df_daily.merge(df_interns[["id","name","department"]], 
                             left_on="intern_id", right_on="id", how="left")
    merged["date"]           = pd.to_datetime(merged["date"]).dt.strftime("%Y-%m-%d")
    merged["arrival_time"]   = pd.to_datetime(merged["arrival_time"]).dt.strftime("%H:%M").fillna("—")
    merged["departure_time"] = pd.to_datetime(merged["departure_time"]).dt.strftime("%H:%M").fillna("—")
    merged["work_duration"]  = merged["work_duration"].round(2).fillna(0)
    merged = merged.sort_values(["date","name"], ascending=[False, True])

    headers = ["Date", "Stagiaire", "Service", "Arrivée", "Départ",
               "Durée (h)", "Statut Check-in", "Statut Check-out", "Attention"]
    for col_i, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_i, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28

    for row_i, (_, row) in enumerate(merged.iterrows(), 2):
        values = [
            row["date"], row["name"], row["department"],
            row["arrival_time"], row["departure_time"],
            row["work_duration"], row.get("checkin_status","—"),
            row.get("checkout_status","—"),
            "Alerte" if row.get("needs_attention") else "OK"
        ]
        for col_i, val in enumerate(values, 1):
            ws.cell(row=row_i, column=col_i, value=val)

        # Color row by status
        status = row.get("status","")
        bg = STATUS_COLORS.get(status, WHITE if row_i % 2 == 0 else GREY_ROW)
        style_data_row(ws, row_i, len(headers), bg)

    auto_width(ws)
    ws.freeze_panes = "A2"


# ── SHEET 2 — RÉSUMÉ PAR STAGIAIRE ───────────────────────────────────────────

def sheet_resume(wb, df_interns, df_daily):
    ws = wb.create_sheet("Résumé Stagiaires")
    ws.sheet_view.showGridLines = False

    headers = ["Stagiaire", "Service", "Jours Présents", "Jours Absents",
               "Taux Présence (%)", "Durée Totale (h)", "Durée Moy./Jour (h)",
               "Retards", "Départs Anticipés", "Score Risque"]
    for col_i, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_i, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28

    for row_i, (_, intern) in enumerate(df_interns.iterrows(), 2):
        intern_daily = df_daily[df_daily["intern_id"] == intern["id"]]
        total        = len(intern_daily)
        present      = len(intern_daily[intern_daily["status"] != "absent"])
        absent       = total - present
        rate         = round(present / total * 100, 1) if total else 0
        total_h      = round(intern_daily["work_duration"].sum(), 2)
        avg_h        = round(intern_daily["work_duration"].mean(), 2) if present else 0
        retards      = len(intern_daily[intern_daily["checkin_status"] == "late"])
        early        = len(intern_daily[intern_daily["checkout_status"] == "early_checkout"])
        risk         = round(100 - rate, 1)

        values = [intern["name"], intern["department"], present, absent,
                  rate, total_h, avg_h, retards, early, risk]
        for col_i, val in enumerate(values, 1):
            ws.cell(row=row_i, column=col_i, value=val)

        # Color by risk
        bg = GREEN if rate >= 80 else ORANGE if rate >= 60 else "FFCDD2"
        style_data_row(ws, row_i, len(headers), bg)

    auto_width(ws)
    ws.freeze_panes = "A2"


# ── SHEET 3 — RÉSUMÉ PAR DÉPARTEMENT ─────────────────────────────────────────

def sheet_departements(wb, df_interns, df_daily):
    ws = wb.create_sheet("Résumé Départements")
    ws.sheet_view.showGridLines = False

    headers = ["Service", "Nb Stagiaires", "Taux Présence Moy. (%)",
               "Total Heures", "Nb Alertes"]
    for col_i, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_i, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28

    for row_i, dept in enumerate(df_interns["department"].unique(), 2):
        dept_interns = df_interns[df_interns["department"] == dept]["id"].tolist()
        dept_daily   = df_daily[df_daily["intern_id"].isin(dept_interns)]
        nb           = len(dept_interns)
        total        = len(dept_daily)
        present      = len(dept_daily[dept_daily["status"] != "absent"])
        rate         = round(present / total * 100, 1) if total else 0
        total_h      = round(dept_daily["work_duration"].sum(), 2)
        alerts       = len(dept_daily[dept_daily["needs_attention"] == True])

        values = [dept, nb, rate, total_h, alerts]
        for col_i, val in enumerate(values, 1):
            ws.cell(row=row_i, column=col_i, value=val)
        style_data_row(ws, row_i, len(headers), GREY_ROW if row_i % 2 == 0 else WHITE)

    auto_width(ws)

    # Bar chart — taux de présence par service
    if wb["Résumé Départements"].max_row > 2:
        chart = BarChart()
        chart.type   = "col"
        chart.title  = "Taux de présence par service"
        chart.y_axis.title = "%"
        chart.style  = 10
        n = wb["Résumé Départements"].max_row - 1
        data   = Reference(ws, min_col=3, min_row=1, max_row=n + 1)
        cats   = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape  = 4
        ws.add_chart(chart, "G2")


# ── SHEET 4 — ALERTES ────────────────────────────────────────────────────────

def sheet_alertes(wb, df_interns, df_daily):
    ws = wb.create_sheet("Journal des Alertes")
    ws.sheet_view.showGridLines = False

    alerts = df_daily[df_daily["needs_attention"] == True].copy()
    if alerts.empty:
        ws["A1"] = "Aucune alerte détectée"
        return

    alerts = alerts.merge(df_interns[["id","name","department"]],
                          left_on="intern_id", right_on="id", how="left")
    alerts["date"] = pd.to_datetime(alerts["date"]).dt.strftime("%Y-%m-%d")
    alerts = alerts.sort_values("date", ascending=False)

    headers = ["Date", "Stagiaire", "Service", "Statut", "Check-in", "Check-out"]
    for col_i, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_i, value=h)
    style_header_row(ws, 1, len(headers), bg="C62828")
    ws.row_dimensions[1].height = 28

    for row_i, (_, row) in enumerate(alerts.iterrows(), 2):
        values = [row["date"], row["name"], row["department"],
                  row.get("status","—"), row.get("checkin_status","—"),
                  row.get("checkout_status","—")]
        for col_i, val in enumerate(values, 1):
            ws.cell(row=row_i, column=col_i, value=val)
        bg = STATUS_COLORS.get(row.get("status",""), "FFCDD2")
        style_data_row(ws, row_i, len(headers), bg)

    auto_width(ws)
    ws.freeze_panes = "A2"


# ── MAIN EXPORT FUNCTION ──────────────────────────────────────────────────────

def export_to_excel(path: str = EXPORT_PATH, department_id: str = None) -> str:
    """
    Génère le fichier Excel complet.
    department_id=None  → export complet (super_admin)
    department_id=UUID  → export filtré (supervisor)
    Retourne le chemin du fichier créé.
    """
    from openpyxl import Workbook

    df_interns, df_daily = fetch_data(department_id=department_id)

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    # Title sheet
    ws_title = wb.create_sheet("Accueil", 0)
    ws_title.sheet_view.showGridLines = False
    ws_title["B2"] = "CHU — Rapport de Pointage des Stagiaires"
    ws_title["B2"].font = Font(bold=True, size=16, color=BLUE_HEADER)
    ws_title["B3"] = f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws_title["B3"].font = Font(size=11, color="666666")
    ws_title["B4"] = f"Total stagiaires : {len(df_interns)}"
    ws_title["B5"] = f"Total enregistrements : {len(df_daily)}"
    ws_title.column_dimensions["B"].width = 50

    sheet_pointages(wb, df_interns, df_daily)
    sheet_resume(wb, df_interns, df_daily)
    sheet_departements(wb, df_interns, df_daily)
    sheet_alertes(wb, df_interns, df_daily)

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    print(f"✅ Excel exporté → {path}")
    return path


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    export_to_excel()

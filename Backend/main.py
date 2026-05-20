"""
CHU Internship Attendance Platform — main.py
=============================================

CHECK-IN tiers (Morocco local time — UTC+1, fixed offset, no DST):
  before 08:30        → rejected  (window not open)
  08:30 – 09:35       → on_time
  09:36 – 10:10       → late
  after  10:10        → missed_checkin  (recorded, NOT rejected)

CHECK-OUT tiers:
  before 15:00        → early_checkout  (recorded with flag, NOT rejected)
  15:00 – 17:00       → completed
  after  17:00        → missed_checkout (recorded, NOT rejected)

OTHER RULES:
  - Any two scans within 5 minutes         → rejected (double-scan guard)
  - No scan at all by end of day           → auto-marked "absent"
    via POST /admin/mark-absences (run once at ~18:00 via cron or manually)
  - Checked in but no checkout by night    → auto-closed as "missed_checkout"
    via POST /admin/auto-close-checkouts   (run once at ~23:00 via cron)
    work_duration = min(CHECKOUT_CLOSE - arrival_time, STANDARD_WORK_HOURS)

STATUS FIELDS on DailyStatus:
  status          → daily admin verdict (one word, shown on dashboard):
                    "on_time" | "late" | "missed_checkin"
                    | "early_checkout" | "missed_checkout" | "absent"
  checkin_status  → "on_time" | "late" | "missed_checkin" | None
  checkout_status → "completed" | "early_checkout" | "missed_checkout" | None
  needs_attention → Boolean flag the dashboard reads to show an alarm icon.
                    True whenever: missed_checkin, early_checkout,
                    missed_checkout, or absent.

NIGHTLY PIPELINE — POST /admin/run-nightly-pipeline:
  Step 1 → auto-close forgotten checkouts
  Step 2 → mark absent interns
  Step 3 → run ETL + feature engineering
  Step 4 → retrain ML models (if enough data exists, else keep rule-based labels)

Times stored as UTC-naive in SQLite (backward-compatible with existing data).
All comparisons use Morocco local time (UTC+1, fixed offset).
"""
from dotenv import load_dotenv
load_dotenv()

# Merged Upstream (Request) and Stashed (status) changes here
from fastapi import FastAPI, Depends, HTTPException, Request, status, UploadFile, File

from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import inspect, text
from database import (
    SessionLocal,
    Intern,
    DailyStatus,
    Department,
    Admin,
    NotificationSetting,
    NotificationLog,
    engine,
    Rotation,
    Etablissement,
    SpecialiteFilieres,
)
from schemas import (
    LoginRequest,
    InternCreate,
    InternUpdate,
    CreateAdminRequest,
    ChangePasswordRequest,
    RotationCreate,
    DepartmentParentUpdate,
)
from auth import (                        
    get_current_admin,
    require_super_admin,
    verify_password,
    create_access_token,
    hash_password,
    AdminRole
)
from pydantic import BaseModel, computed_field
from typing import Optional, List
from datetime import datetime, time, timezone, timedelta, date as date_type
import uuid
import subprocess
import os
from fastapi import Body
from fastapi.responses import FileResponse, StreamingResponse
import tempfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from apscheduler.schedulers.background import BackgroundScheduler
from contextlib import asynccontextmanager
import json
import smtplib
from email.message import EmailMessage

from excel_export import export_to_excel
import threading

def export_in_background():
    threading.Thread(target=export_to_excel, daemon=True).start()


def _ensure_intern_rotation_columns():
    """Add lightweight rotation planning columns on existing databases."""
    inspector = inspect(engine)
    existing = {col["name"] for col in inspector.get_columns("interns")}
    missing = [name for name in ("start_date", "end_date", "type", "school", "archived") if name not in existing]
    if not missing:
        return
    with engine.begin() as conn:
        for column in missing:
            column_type = "BOOLEAN" if column == "archived" else "VARCHAR"
            conn.execute(text(f"ALTER TABLE interns ADD COLUMN {column} {column_type}"))
    print(f"Intern rotation columns added: {', '.join(missing)}")


def _ensure_admin_email_column():
    """Add email column to existing admin tables."""
    inspector = inspect(engine)
    existing = {col["name"] for col in inspector.get_columns("admins")}
    if "email" in existing:
        return
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE admins ADD COLUMN email VARCHAR"))
    print("Admin email column added")


def _parse_optional_iso_date(value: str | None, field_name: str) -> str | None:
    if value in (None, ""):
        return None
    try:
        return date_type.fromisoformat(value).isoformat()
    except ValueError:
        raise HTTPException(status_code=400, detail=f"{field_name} doit etre une date valide au format YYYY-MM-DD")


def _validate_rotation_dates(start_date: str | None, end_date: str | None):
    if start_date and end_date and date_type.fromisoformat(start_date) > date_type.fromisoformat(end_date):
        raise HTTPException(status_code=400, detail="La date de debut doit etre avant ou egale a la date de fin")


# ── APSCHEDULER SETUP ──────────────────────────────────────────────────────────
# Nightly workflow scheduler — runs Mon–Fri only

class NotificationSettingsPayload(BaseModel):
    enabled: bool = False
    smtp_host: str | None = None
    smtp_port: int | None = 587
    smtp_username: str | None = None
    smtp_password: str | None = None
    smtp_from_email: str | None = None
    absence_days: int = 3
    stage_end_hours: int = 48


scheduler = BackgroundScheduler()


def _should_run_today() -> bool:
    """Return True only if today is Mon–Fri (0=Mon, 6=Sun)."""
    return now_local().weekday() < 5  # 0-4 = Mon-Fri


def _get_notification_settings(db: Session) -> NotificationSetting:
    settings = db.query(NotificationSetting).filter(NotificationSetting.id == "default").first()
    if settings:
        return settings
    settings = NotificationSetting(id="default", enabled=False, channel="email", absence_days=3, stage_end_hours=48)
    db.add(settings)
    db.commit()
    db.refresh(settings)
    return settings


def _notification_settings_dict(settings: NotificationSetting, include_secret: bool = False) -> dict:
    return {
        "enabled": bool(settings.enabled),
        "channel": "email",
        "smtp_host": settings.smtp_host or "",
        "smtp_port": settings.smtp_port or 587,
        "smtp_username": settings.smtp_username or "",
        "smtp_password": settings.smtp_password if include_secret else "",
        "smtp_from_email": settings.smtp_from_email or "",
        "absence_days": settings.absence_days or 3,
        "stage_end_hours": settings.stage_end_hours or 48,
    }


def _working_days_ending(target_date: date_type, count: int) -> list[date_type]:
    days = []
    cursor = target_date
    while len(days) < count:
        if cursor.weekday() < 5:
            days.append(cursor)
        cursor -= timedelta(days=1)
    return list(reversed(days))


def _has_status_on_date(db: Session, intern_id: str, target_date: date_type, status_value: str) -> bool:
    day_start = datetime(target_date.year, target_date.month, target_date.day)
    day_end = day_start + timedelta(days=1)
    return db.query(DailyStatus.id).filter(
        DailyStatus.intern_id == intern_id,
        DailyStatus.date >= day_start,
        DailyStatus.date < day_end,
        DailyStatus.status == status_value,
    ).first() is not None


def _department_name_map(db: Session) -> dict[str, str]:
    return {d.id: d.name for d in db.query(Department).all()}


def _admin_email(admin: Admin | None) -> str | None:
    if not admin:
        return None
    if getattr(admin, "email", None):
        return admin.email
    return admin.username if "@" in (admin.username or "") else None


def _dfri_email(db: Session) -> str | None:
    admin = db.query(Admin).filter(Admin.role == AdminRole.DFRI.value, Admin.email != None).first()  # noqa: E711
    if admin:
        return admin.email
    admin = db.query(Admin).filter(Admin.role == AdminRole.DFRI.value).first()
    return _admin_email(admin)


def _chef_service_email(db: Session, department_id: str | None) -> str | None:
    if not department_id:
        return None
    admin = db.query(Admin).filter(
        Admin.role == AdminRole.CHEF_SERVICE.value,
        Admin.department_id == department_id,
        Admin.email != None,  # noqa: E711
    ).first()
    if admin:
        return admin.email
    admin = db.query(Admin).filter(
        Admin.role == AdminRole.CHEF_SERVICE.value,
        Admin.department_id == department_id,
    ).first()
    return _admin_email(admin)


def _collect_notification_events(db: Session, target_date: date_type | None = None) -> list[dict]:
    target_date = target_date or now_local().date()
    settings = _get_notification_settings(db)
    absence_days = max(int(settings.absence_days or 3), 1)
    stage_end_hours = max(int(settings.stage_end_hours or 48), 1)
    departments = _department_name_map(db)
    interns = db.query(Intern).filter((Intern.archived == False) | (Intern.archived == None)).all()  # noqa: E712,E711
    events = []

    absence_window = _working_days_ending(target_date, absence_days)
    for intern in interns:
        full_name = f"{intern.first_name} {intern.last_name}".strip()
        dept_name = departments.get(intern.department_id, "Service inconnu")
        if all(_has_status_on_date(db, intern.id, day, "absent") for day in absence_window):
            recipient = _chef_service_email(db, intern.department_id)
            events.append({
                "event_type": "absence_consecutive",
                "intern_id": intern.id,
                "department_id": intern.department_id,
                "recipient": recipient,
                "dedupe_key": f"absence_consecutive:{intern.id}:{absence_window[-1]}:{absence_days}",
                "title": f"Absence consecutive - {full_name}",
                "message": f"{full_name} est absent depuis {absence_days} jours ouvrables consecutifs dans le service {dept_name}. Cette alerte est destinee au chef de service.",
            })

        if intern.end_date:
            try:
                end_date = date_type.fromisoformat(intern.end_date)
            except ValueError:
                end_date = None
            if end_date:
                hours_left = (datetime(end_date.year, end_date.month, end_date.day, 23, 59, tzinfo=TZ) - now_local()).total_seconds() / 3600
                if 0 <= hours_left <= stage_end_hours:
                    recipient = _dfri_email(db)
                    events.append({
                        "event_type": "stage_ending",
                        "intern_id": intern.id,
                        "department_id": intern.department_id,
                        "recipient": recipient,
                        "dedupe_key": f"stage_ending:{intern.id}:{intern.end_date}",
                        "title": f"Stage a confirmer - {full_name}",
                        "message": f"Le stage de {full_name} se termine le {intern.end_date} dans le service {dept_name}. Verifiez la prolongation ou l'archivage.",
                    })

        if intern.start_date:
            try:
                start_date = date_type.fromisoformat(intern.start_date)
            except ValueError:
                start_date = None
            if start_date == target_date + timedelta(days=1):
                recipient = _chef_service_email(db, intern.department_id)
                events.append({
                    "event_type": "rotation_tomorrow",
                    "intern_id": intern.id,
                    "department_id": intern.department_id,
                    "recipient": recipient,
                    "dedupe_key": f"rotation_tomorrow:{intern.id}:{intern.start_date}",
                    "title": f"Rotation demain - {full_name}",
                    "message": f"{full_name} commence une rotation demain ({intern.start_date}) dans le service {dept_name}.",
                })

    return events


def _send_email_notification(settings: NotificationSetting, recipient_email: str | None, title: str, message: str):
    if not recipient_email:
        raise RuntimeError("Aucun email destinataire configure")
    if not settings.smtp_host:
        raise RuntimeError("Serveur SMTP non configure")

    email = EmailMessage()
    email["Subject"] = title
    email["From"] = settings.smtp_from_email or settings.smtp_username or recipient_email
    email["To"] = recipient_email
    email.set_content(message)

    port = int(settings.smtp_port or 587)
    with smtplib.SMTP(settings.smtp_host, port, timeout=15) as smtp:
        smtp.starttls()
        if settings.smtp_username:
            smtp.login(settings.smtp_username, settings.smtp_password or "")
        smtp.send_message(email)


def _dispatch_notification(settings: NotificationSetting, recipient_email: str | None, title: str, message: str) -> tuple[str, str | None, str]:
    try:
        _send_email_notification(settings, recipient_email, title, message)
        return "sent", None, recipient_email or ""
    except Exception as exc:
        return "error", str(exc), recipient_email or ""


def _run_notification_job_for_date(target_date: date_type, db: Session, force_send: bool = False) -> dict:
    settings = _get_notification_settings(db)
    events = _collect_notification_events(db, target_date)
    result = {"date": str(target_date), "detected": len(events), "sent": 0, "skipped": 0, "errors": 0}

    if not settings.enabled and not force_send:
        return {**result, "message": "Notifications desactivees"}

    for event in events:
        existing = db.query(NotificationLog).filter(NotificationLog.dedupe_key == event["dedupe_key"]).first()
        if existing and not force_send:
            result["skipped"] += 1
            continue

        if force_send:
            event["dedupe_key"] = f"{event['dedupe_key']}:manual:{uuid.uuid4()}"

        status_value, error, recipient = _dispatch_notification(settings, event.get("recipient"), event["title"], event["message"])
        db.add(NotificationLog(
            id=str(uuid.uuid4()),
            event_type=event["event_type"],
            intern_id=event["intern_id"],
            department_id=event["department_id"],
            dedupe_key=event["dedupe_key"],
            title=event["title"],
            message=event["message"],
            channel="email",
            recipient=recipient,
            status=status_value,
            error=error,
            created_at=datetime.utcnow(),
        ))
        if status_value == "error":
            result["errors"] += 1
        else:
            result["sent"] += 1

    db.commit()
    return result


def _auto_close_checkouts_for_date(target_date: date_type, db: Session) -> dict:
    """
    Auto-close unclosed checkins for an arbitrary target_date.
    Used both by the live endpoint (today) and the catch-up pipeline (past dates).
    """
    day_start = datetime(target_date.year, target_date.month, target_date.day)
    day_end   = day_start + timedelta(days=1)

    unclosed = db.query(DailyStatus).filter(
        DailyStatus.date >= day_start,
        DailyStatus.date <  day_end,
        DailyStatus.arrival_time   != None,   # noqa: E711
        DailyStatus.departure_time == None,   # noqa: E711
    ).all()

    closed_count = 0
    # Build a timezone-aware reference for CHECKOUT_CLOSE on target_date
    close_dt = datetime(
        target_date.year, target_date.month, target_date.day,
        CHECKOUT_CLOSE.hour, CHECKOUT_CLOSE.minute, 0,
        tzinfo=TZ,
    )

    for daily in unclosed:
        arr_local      = to_local(daily.arrival_time)
        credited_hours = round(
            min((close_dt - arr_local).total_seconds() / 3600, STANDARD_WORK_HOURS),
            2,
        )
        credited_hours = max(credited_hours, 0.0)

        daily.departure_time  = to_utc_naive(close_dt)
        daily.work_duration   = credited_hours
        daily.checkout_status = "missed_checkout"
        daily.needs_attention = True
        if daily.status != "missed_checkin":
            daily.status = "missed_checkout"

        closed_count += 1

    db.commit()
    return {
        "message": f"{closed_count} checkout(s) fermé(s) pour {target_date}",
        "date": str(target_date),
        "count": closed_count,
    }


def _mark_absences_for_date(target_date: date_type, db: Session) -> dict:
    """
    Mark absent interns for an arbitrary target_date.
    Used both by the live endpoint (today) and the catch-up pipeline (past dates).
    """
    day_start = datetime(target_date.year, target_date.month, target_date.day)
    day_end   = day_start + timedelta(days=1)

    scanned = {
        row.intern_id
        for row in db.query(DailyStatus.intern_id).filter(
            DailyStatus.date >= day_start,
            DailyStatus.date <  day_end,
        ).all()
    }

    # Use noon on target_date as the stored timestamp (neutral, avoids timezone edge cases)
    record_dt = datetime(
        target_date.year, target_date.month, target_date.day, 12, 0, 0,
        tzinfo=TZ,
    )

    absent_count = 0
    for intern in db.query(Intern).all():
        if intern.id not in scanned:
            db.add(DailyStatus(
                id=str(uuid.uuid4()),
                intern_id=intern.id,
                date=to_utc_naive(record_dt),
                arrival_time=None,
                departure_time=None,
                status="absent",
                checkin_status=None,
                checkout_status=None,
                work_duration=0.0,
                needs_attention=True,
            ))
            absent_count += 1

    db.commit()
    return {
        "message": f"{absent_count} stagiaire(s) absent(s) pour {target_date}",
        "date": str(target_date),
        "count": absent_count,
    }


def _run_auto_close_job():
    """Scheduled job: Auto-close forgotten checkouts at 23:00 on weekdays."""
    if not _should_run_today():
        return
    db = SessionLocal()
    try:
        today = now_local().date()
        _auto_close_checkouts_for_date(today, db)
    finally:
        db.close()


def _run_mark_absences_job():
    """Scheduled job: Mark absent interns at 23:05 on weekdays."""
    if not _should_run_today():
        return
    db = SessionLocal()
    try:
        today = now_local().date()
        _mark_absences_for_date(today, db)
    finally:
        db.close()


def _run_notifications_job():
    """Scheduled job: send configured automatic alerts after attendance cleanup."""
    if not _should_run_today():
        return
    db = SessionLocal()
    try:
        _run_notification_job_for_date(now_local().date(), db)
    finally:
        db.close()


def _run_etl_ml_job():
    """Scheduled job: Run ETL + ML retraining at 23:30 on weekdays. Records run date."""
    if not _should_run_today():
        return
    # ── ETL ───────────────────────────────────────────────────────────
    try:
        etl_result = subprocess.run(
            ["python", "etl/transform.py"],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if etl_result.returncode != 0:
            print(f"ETL error: {etl_result.stderr}")
    except Exception as e:
        print(f"ETL job failed: {e}")

    # ── ML retrain ────────────────────────────────────────────────────
    ml_script = os.path.join("ml", "train.py")
    if os.path.exists(ml_script):
        try:
            ml_result = subprocess.run(
                ["python", ml_script],
                capture_output=True,
                text=True,
                timeout=120,
            )
            if ml_result.returncode != 0:
                print(f"ML training error: {ml_result.stderr}")
        except Exception as e:
            print(f"ML training job failed: {e}")

    try:
        export_path = export_to_excel()
        print(f"Nightly Excel export refreshed: {export_path}")
    except Exception as e:
        print(f"Nightly Excel export failed: {e}")

    # ── Record that the pipeline ran today ────────────────────────────
    _write_last_pipeline_run(now_local().date())

    # Archive interns whose end_date has passed
    db = SessionLocal()
    try:
        today_str = str(now_local().date())
        expired = db.query(Intern).filter(
            Intern.end_date != None,
            Intern.end_date < today_str,
            Intern.archive_status == "active",
        ).all()
        for intern in expired:
            intern.archive_status = "pending_archive"
        db.commit()
        if expired:
            print(f"Nightly: {len(expired)} intern(s) marked pending_archive")
    finally:
        db.close()
    print(f"Nightly pipeline completed and recorded for {now_local().date()}")


# ── PIPELINE LAST-RUN TRACKER ─────────────────────────────────────────────────
# We store the last date the nightly pipeline ran in a small file next to
# main.py. This survives server restarts and is much simpler than a DB table.

PIPELINE_LAST_RUN_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".pipeline_last_run")


def _read_last_pipeline_run() -> date_type | None:
    """Return the date stored in .pipeline_last_run, or None if not found."""
    try:
        with open(PIPELINE_LAST_RUN_FILE, "r") as f:
            return date_type.fromisoformat(f.read().strip())
    except Exception:
        return None


def _write_last_pipeline_run(d: date_type):
    """Persist the pipeline run date to disk."""
    try:
        with open(PIPELINE_LAST_RUN_FILE, "w") as f:
            f.write(str(d))
    except Exception as e:
        print(f"Could not write pipeline last-run file: {e}")


def _get_last_working_day(reference: date_type) -> date_type:
    """
    Return the most recent weekday strictly before `reference`.
    Monday → previous Friday (skips Saturday and Sunday).
    """
    day = reference - timedelta(days=1)
    while day.weekday() >= 5:   # 5=Sat, 6=Sun
        day -= timedelta(days=1)
    return day


def _run_pipeline_for_date(target_date: date_type, db: Session) -> dict:
    """
    Run steps 1 and 2 of the nightly pipeline (close checkouts + mark absences)
    for an arbitrary past date, then run ETL + ML retraining.
    Steps 1 & 2 are date-aware; steps 3 & 4 always operate on the full dataset.
    """
    results = {}

    # Step 1: Auto-close forgotten checkouts for target_date
    try:
        results["step1_auto_close"] = _auto_close_checkouts_for_date(target_date, db)
    except Exception as e:
        results["step1_auto_close"] = {"error": str(e)}

    # Step 2: Mark absent interns for target_date
    try:
        results["step2_mark_absences"] = _mark_absences_for_date(target_date, db)
    except Exception as e:
        results["step2_mark_absences"] = {"error": str(e)}

    # Step 3: Send automatic notifications
    try:
        results["step3_notifications"] = _run_notification_job_for_date(target_date, db)
    except Exception as e:
        results["step3_notifications"] = {"error": str(e)}

    # Step 3: ETL
    try:
        etl_result = subprocess.run(
            ["python", "etl/transform.py"],
            capture_output=True, text=True, timeout=120,
        )
        results["step3_etl"] = {
            "returncode": etl_result.returncode,
            "stdout": etl_result.stdout[-500:] if etl_result.stdout else "",
            "stderr": etl_result.stderr[-300:] if etl_result.stderr else "",
        }
    except Exception as e:
        results["step3_etl"] = {"error": str(e)}

    # Step 4: ML retrain
    ml_script = os.path.join("ml", "train.py")
    if os.path.exists(ml_script):
        try:
            ml_result = subprocess.run(
                ["python", ml_script],
                capture_output=True, text=True, timeout=120,
            )
            results["step4_ml_retrain"] = {
                "returncode": ml_result.returncode,
                "stdout": ml_result.stdout[-500:] if ml_result.stdout else "",
                "stderr": ml_result.stderr[-300:] if ml_result.stderr else "",
            }
        except Exception as e:
            results["step4_ml_retrain"] = {"error": str(e)}
    else:
        results["step4_ml_retrain"] = {
            "skipped": True,
            "reason": "ml/train.py not found yet — will run once ML module is added",
        }

    try:
        results["step5_excel_export"] = {"path": export_to_excel()}
    except Exception as e:
        results["step5_excel_export"] = {"error": str(e)}

    return results


def _check_and_run_missed_pipeline(db: Session):
    """
    On server startup, check whether the nightly pipeline ran for the last
    working day. If it didn't (server was down overnight), run it now so no
    day is ever skipped.

    Logic:
      - Read the last pipeline run date from .pipeline_last_run.
      - Compute the last weekday before today.
      - If last_run < last_working_day → pipeline was missed → run it for
        last_working_day and record the date.
      - If already up-to-date → do nothing.
    """
    today            = now_local().date()
    last_working_day = _get_last_working_day(today)
    last_run         = _read_last_pipeline_run()

    # Nothing to catch up if we already ran for the last working day (or later)
    if last_run is not None and last_run >= last_working_day:
        print(f"Pipeline up-to-date (last run: {last_run}, last working day: {last_working_day})")
        return False

    # Also skip on weekends when there are literally no working days to catch up
    # (e.g. server restarted on a Saturday morning right after Friday's pipeline ran)
    if last_run == last_working_day:
        print(f"Pipeline already ran for {last_working_day}")
        return False

    print(f"CATCH-UP: Pipeline missed for {last_working_day} — running now...")
    try:
        results = _run_pipeline_for_date(last_working_day, db)
        _write_last_pipeline_run(last_working_day)
        print(f"CATCH-UP: Pipeline completed for {last_working_day}. Results: {results}")
        return True
    except Exception as e:
        print(f"CATCH-UP: Pipeline failed for {last_working_day}: {e}")
        return False


@asynccontextmanager
async def lifespan(app: FastAPI):
    """FastAPI lifespan: start scheduler on startup, stop on shutdown."""
    # ── Startup ───────────────────────────────────────────────────────
    _ensure_intern_rotation_columns()
    _ensure_admin_email_column()
    scheduler.add_job(_run_auto_close_job, "cron", hour=23, minute=0, id="auto_close")
    scheduler.add_job(_run_mark_absences_job, "cron", hour=23, minute=5, id="mark_absences")
    scheduler.add_job(_run_notifications_job, "cron", hour=23, minute=10, id="notifications")
    scheduler.add_job(_run_etl_ml_job, "cron", hour=23, minute=30, id="etl_ml")
    scheduler.start()
    print("APScheduler started - nightly jobs scheduled for 23:00, 23:05, 23:10, 23:30")

    # ── CATCH-UP: Check if nightly pipeline was missed ────────────────
    db = SessionLocal()
    try:
        _check_and_run_missed_pipeline(db)
    finally:
        db.close()

    yield
    # ── Shutdown ──────────────────────────────────────────────────────
    scheduler.shutdown(wait=True)
    print("APScheduler stopped")

# ── TIMEZONE ──────────────────────────────────────────────────────────────────
# Morocco has been permanently on UTC+1 since October 2018 (no DST).
TZ = timezone(timedelta(hours=1), name="Morocco/UTC+1")


def now_local() -> datetime:
    """Return the current datetime in Morocco local time (UTC+1, aware)."""
    return datetime.now(TZ)


def to_local(dt: datetime) -> datetime:
    """Convert a naive UTC datetime (from SQLite) → Morocco local time (UTC+1)."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(TZ)


def to_utc_naive(dt: datetime) -> datetime:
    """Convert a Morocco-aware datetime → UTC-naive for SQLite storage."""
    return dt.astimezone(timezone.utc).replace(tzinfo=None)


# ── TIME WINDOWS ──────────────────────────────────────────────────────────────
CHECKIN_OPEN  = time(8,  30)
CHECKIN_LATE  = time(9,  35)
CHECKIN_CLOSE = time(10, 10)

CHECKOUT_OPEN  = time(15, 0)
CHECKOUT_CLOSE = time(17, 0)

# Standard credited work hours when checkout is auto-closed at night
# An intern who arrived on time (08:30) and forgot checkout gets credited
# min(CHECKOUT_CLOSE - arrival_time, STANDARD_WORK_HOURS)
STANDARD_WORK_HOURS = 8.0

# Minimum gap between any two scans for the same intern
DOUBLE_SCAN_MINUTES = 5

# Statuses that should raise an alarm on the admin dashboard
ATTENTION_STATUSES = {"missed_checkin", "early_checkout", "missed_checkout", "absent"}

# Minimum number of intern-days of data before ML models are trusted
# Below this threshold the system falls back to rule-based labels
ML_MIN_SAMPLES = 30

ARCHIVE_STATUS_ACTIVE = "active"
ARCHIVE_STATUS_PENDING = "pending_archive"
ARCHIVE_STATUS_ARCHIVED = "archived"
INACTIVE_STAGE_STATUSES = {ARCHIVE_STATUS_PENDING, ARCHIVE_STATUS_ARCHIVED}

FILIERE_MEDICINE = "medicine"
FILIERE_NURSE = "nurse"
FILIERE_TECH_SANTE_LAB_BIOLOGY = "tech_sante_lab_biology"
FILIERE_ADMINISTRATIVE = "administrative"

FILIERE_ALIASES = {
    FILIERE_MEDICINE: {FILIERE_MEDICINE, "Médicale"},
    FILIERE_NURSE: {FILIERE_NURSE, "Infirmière"},
    FILIERE_TECH_SANTE_LAB_BIOLOGY: {FILIERE_TECH_SANTE_LAB_BIOLOGY, "Techniques Santé/Lab"},
    FILIERE_ADMINISTRATIVE: {FILIERE_ADMINISTRATIVE, "Administrative"},
}
VALID_FILIERES = set(FILIERE_ALIASES.keys()).union(*FILIERE_ALIASES.values())
SPECIALITE_REQUIRED_FILIERES = {FILIERE_MEDICINE}
SPECIALITE_ALLOWED_FILIERES = {FILIERE_MEDICINE, FILIERE_TECH_SANTE_LAB_BIOLOGY}
YEAR_ALLOWED_FILIERES = {FILIERE_MEDICINE, FILIERE_NURSE}


def _normalize_filiere(value: str | None) -> str | None:
    if value in (None, ""):
        return None
    raw = str(getattr(value, "value", value))
    for official, aliases in FILIERE_ALIASES.items():
        if raw in aliases:
            return official
    raise HTTPException(status_code=400, detail="Filiere invalide")


def _filiere_filter_values(value: str) -> list[str]:
    official = _normalize_filiere(value)
    return sorted(FILIERE_ALIASES[official])


def _validate_intern_business_rules(
    filiere: str | None,
    etablissement_id: str | None,
    specialite_id: str | None,
    annee: int | None,
) -> None:
    official = _normalize_filiere(filiere)
    if not official:
        return
    if not etablissement_id:
        raise HTTPException(status_code=400, detail="Etablissement requis pour cette filiere")
    if official in SPECIALITE_REQUIRED_FILIERES and not specialite_id:
        raise HTTPException(status_code=400, detail="Specialite requise pour la filiere medicine")
    if specialite_id and official not in SPECIALITE_ALLOWED_FILIERES:
        raise HTTPException(status_code=400, detail="Specialite non applicable pour cette filiere")
    if annee is not None and official not in YEAR_ALLOWED_FILIERES:
        raise HTTPException(status_code=400, detail="Annee/promotion non applicable pour cette filiere")


def _validate_intern_filter_rules(
    filiere: str | None,
    specialite_id: str | None,
    annee: int | None,
) -> None:
    official = _normalize_filiere(filiere)
    if not official:
        return
    if specialite_id and official not in SPECIALITE_ALLOWED_FILIERES:
        raise HTTPException(status_code=400, detail="Specialite non applicable pour cette filiere")
    if annee is not None and official not in YEAR_ALLOWED_FILIERES:
        raise HTTPException(status_code=400, detail="Annee/promotion non applicable pour cette filiere")


# ── CHECKIN STATUS RESOLVER ───────────────────────────────────────────────────

def resolve_checkin_status(t: time) -> str:
    if t < CHECKIN_OPEN:
        raise HTTPException(
            status_code=400,
            detail="Too early for check-in — window opens at 08:30",
        )
    if t <= CHECKIN_LATE:
        return "on_time"
    if t <= CHECKIN_CLOSE:
        return "late"
    return "missed_checkin"


# ── CHECKOUT STATUS RESOLVER ──────────────────────────────────────────────────

def resolve_checkout_status(t: time) -> str:
    if t < CHECKOUT_OPEN:
        return "early_checkout"
    if t <= CHECKOUT_CLOSE:
        return "completed"
    return "missed_checkout"


# ── DOUBLE-SCAN GUARD ─────────────────────────────────────────────────────────

def check_double_scan(daily: DailyStatus, now: datetime) -> None:
    last_scan = daily.departure_time or daily.arrival_time
    if last_scan is None:
        return
    gap_minutes = (now - to_local(last_scan)).total_seconds() / 60
    if gap_minutes < DOUBLE_SCAN_MINUTES:
        remaining = int(DOUBLE_SCAN_MINUTES - gap_minutes)
        raise HTTPException(
            status_code=429,
            detail=(
                f"Scan ignored: too soon after previous scan — "
                f"wait {remaining} more minute(s)"
            ),
        )


# ── DB HELPER ─────────────────────────────────────────────────────────────────

def get_today_record(intern_id: str, today: date_type, db: Session):
    """Return today's DailyStatus for an intern, or None if no scan yet."""
    return db.query(DailyStatus).filter(
        DailyStatus.intern_id == intern_id,
        DailyStatus.date >= datetime(today.year, today.month, today.day),
        DailyStatus.date <  datetime(today.year, today.month, today.day) + timedelta(days=1),
    ).first()


def get_unclosed_checkins(today: date_type, db: Session):
    """
    Return all DailyStatus rows for today where the intern checked in
    but never checked out. These are candidates for auto-close.
    """
    return db.query(DailyStatus).filter(
        DailyStatus.date >= datetime(today.year, today.month, today.day),
        DailyStatus.date <  datetime(today.year, today.month, today.day, 23, 59, 59),
        DailyStatus.arrival_time  != None,   # noqa: E711 — SQLAlchemy needs !=
        DailyStatus.departure_time == None,  # noqa: E711
    ).all()


# ── PYDANTIC SCHEMAS ──────────────────────────────────────────────────────────

class DailyStatusOut(BaseModel):
    id: str
    intern_id: str
    status: Optional[str]
    checkin_status: Optional[str]
    checkout_status: Optional[str]
    needs_attention: Optional[bool]
    work_duration: Optional[float]
    date: Optional[datetime]
    arrival_time: Optional[datetime]
    departure_time: Optional[datetime]

    model_config = {"from_attributes": True}

    def model_post_init(self, __context):
        if self.date:
            self.date = to_local(self.date)
        if self.arrival_time:
            self.arrival_time = to_local(self.arrival_time)
        if self.departure_time:
            self.departure_time = to_local(self.departure_time)


class ScanRequest(BaseModel):
    intern_id: str


"""class InternCreate(BaseModel):
    first_name: str
    last_name: str
    department_id: str
    start_date: str | None = None
    end_date: str | None = None"""


# ── APP SETUP ─────────────────────────────────────────────────────────────────

app = FastAPI(title="CHU Plateforme de Pointage", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def add_no_cache_headers(request: Request, call_next):
    response = await call_next(request)
    public_routes = ["/", "/auth/login", "/docs", "/openapi.json", "/static"]
    is_public = any(request.url.path.startswith(r) for r in public_routes)
    if not is_public:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, private"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ── /scan ─────────────────────────────────────────────────────────────────────

@app.post("/scan")
def register_scan(request: ScanRequest, db: Session = Depends(get_db)):

    intern = db.query(Intern).filter(Intern.id == request.intern_id).first()
    if not intern:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    if intern.archive_status in INACTIVE_STAGE_STATUSES:
        return {
            "event": "inactive_stage",
            "message": "Stage terminé",
        }

    now   = now_local()
    today = now.date()
    t     = now.time()

    daily = get_today_record(request.intern_id, today, db)

    # ── Lookup active rotation for today ──────────────────────────────────
    today_str = today.isoformat()
    active_rotation = db.query(Rotation).filter(
        Rotation.intern_id == request.intern_id,
        Rotation.date_debut <= today_str,
        Rotation.date_fin >= today_str,
    ).first()
    effective_dept_id = active_rotation.department_id if active_rotation else intern.department_id

    # ── CASE A: No record today → CHECK-IN ───────────────────────────────
    if daily is None:
        checkin_status  = resolve_checkin_status(t)
        needs_attention = checkin_status in ATTENTION_STATUSES

        daily = DailyStatus(
            id=str(uuid.uuid4()),
            intern_id=request.intern_id,
            department_id=effective_dept_id,
            date=to_utc_naive(now),
            arrival_time=to_utc_naive(now),
            departure_time=None,
            status=checkin_status,
            checkin_status=checkin_status,
            checkout_status=None,
            work_duration=0.0,
            needs_attention=needs_attention,
        )
        db.add(daily)
        db.commit()
        

        labels = {
            "on_time":        "À l'heure",
            "late":           "⏰ En retard",
            "missed_checkin": "Hors créneau — enregistré comme missed check-in",
        }

       
        export_in_background()   # ← la nouvelle ligne
    
        return {
            "event":           "check_in",
            "checkin_status":  checkin_status,
            "needs_attention": needs_attention,
            "message": (
                f"Bonjour {intern.first_name} ! "
                f"Arrivée à {now.strftime('%H:%M')} — {labels[checkin_status]}"
            ),
        }

    # ── Double-scan guard ─────────────────────────────────────────────────
    check_double_scan(daily, now)

    # ── CASE B: Checked in, no checkout yet → CHECK-OUT ──────────────────
    if daily.arrival_time and not daily.departure_time:

        checkout_status = resolve_checkout_status(t)
        arr_local       = to_local(daily.arrival_time)
        duration        = round((now - arr_local).total_seconds() / 3600, 2)

        daily.departure_time  = to_utc_naive(now)
        daily.work_duration   = duration
        daily.checkout_status = checkout_status

        if checkout_status in ATTENTION_STATUSES:
            daily.status          = checkout_status
            daily.needs_attention = True

        db.commit()

        labels = {
            "completed":       f"Départ à {now.strftime('%H:%M')} — {duration}h travaillées",
            "early_checkout":  f"Départ anticipé à {now.strftime('%H:%M')} — {duration}h (visible sur dashboard)",
            "missed_checkout": f"🔴 Hors créneau à {now.strftime('%H:%M')} — {duration}h travaillées",
        }

        export_in_background()
        
        return {
            "event":           "check_out",
            "checkout_status": checkout_status,
            "needs_attention": checkout_status in ATTENTION_STATUSES,
            "work_duration":   duration,
            "message":         f"Au revoir {intern.first_name} ! {labels[checkout_status]}",
        }

    # ── CASE C: Both already recorded ────────────────────────────────────
    return {
        "event":   "already_complete",
        "message": f"Pointage déjà complet pour aujourd'hui ({intern.first_name})",
    }


# ── /admin/auto-close-checkouts ───────────────────────────────────────────────

@app.post("/admin/auto-close-checkouts", tags=["Administration"])
def auto_close_checkouts(db: Session = Depends(get_db)):
    """
    Auto-close any intern who checked in today but never checked out.
    Delegates to the date-aware helper using today's date.
    """
    return _auto_close_checkouts_for_date(now_local().date(), db)

# ── DEPARTMENT MANAGEMENT ─────────────────────────────────────────────────────
 
@app.post("/departments", tags=["Administration"])
def create_department(
    name: str = Body(..., embed=True),
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_super_admin),
):
    """Create a new department (super_admin only)."""
    existing = db.query(Department).filter(Department.name == name.strip()).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"Le service '{name}' existe déjà.")
    dept = Department(id=str(uuid.uuid4()), name=name.strip())
    db.add(dept)
    db.commit()
    db.refresh(dept)
    return {"id": dept.id, "name": dept.name}
 
 
@app.delete("/departments/{dept_id}", tags=["Administration"])
def delete_department(
    dept_id: str,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_super_admin),
):
    """Delete a department — blocked if interns are still assigned to it."""
    dept = db.query(Department).filter(Department.id == dept_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Service introuvable.")
    linked = db.query(Intern).filter(Intern.department_id == dept_id).count()
    if linked > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Impossible de supprimer : {linked} stagiaire(s) sont encore assignés à ce service.",
        )
    db.delete(dept)
    db.commit()
    return {"message": f"Service '{dept.name}' supprimé."}
 
 
# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────
 
@app.get("/export/excel", tags=["Export"])
def export_excel(
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin),
):
    """
    Generate and return the Excel report.
    Super-admins get the full file; supervisors get their department only.
    """
    from excel_export import export_to_excel
    import shutil

    dept_id = None if admin.role == AdminRole.DFRI.value else admin.department_id

    # Write to a named temp file that persists until the OS cleans it up.
    # We must NOT use TemporaryDirectory here — it deletes the file before
    # FileResponse can stream it, which caused the "fail to fetch" error.
    fd, stable_path = tempfile.mkstemp(suffix=".xlsx", prefix="CHU_Pointages_")
    os.close(fd)  # close the raw file descriptor; export_to_excel opens it itself

    try:
        export_to_excel(path=stable_path, department_id=dept_id)
    except Exception as e:
        os.unlink(stable_path)
        raise HTTPException(status_code=500, detail=f"Erreur génération Excel : {e}")

    return FileResponse(
        stable_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="CHU_Pointages.xlsx",
        background=None,  # FileResponse will stream then the OS temp cleaner handles it
    )
 

# ── /admin/mark-absences ──────────────────────────────────────────────────────

def _mark_absences_logic(db: Session) -> dict:
    """Core absence-marking logic for today. Delegates to the date-aware helper."""
    return _mark_absences_for_date(now_local().date(), db)

@app.post("/admin/mark-absences", tags=["Administration"])
def mark_absences(db: Session = Depends(get_db), admin: Admin = Depends(require_super_admin)):
    return _mark_absences_logic(db)

@app.delete("/admin/undo-absences", tags=["Administration"])
def undo_absences(db: Session = Depends(get_db), admin: Admin = Depends(require_super_admin)):
    """Delete all 'absent' DailyStatus rows created today (undo accidental mark-absences)."""
    now   = now_local()
    today = now.date()
    deleted = db.query(DailyStatus).filter(
        DailyStatus.date >= datetime(today.year, today.month, today.day),
        DailyStatus.date <  datetime(today.year, today.month, today.day) + timedelta(days=1),
        DailyStatus.status == "absent",
    ).delete(synchronize_session=False)
    db.commit()
    return {"message": f"✅ {deleted} absence(s) supprimée(s) pour aujourd'hui ({today})", "count": deleted}

# ── /admin/run-nightly-pipeline ───────────────────────────────────────────────

@app.post("/admin/run-nightly-pipeline", tags=["Administration"])
def run_nightly_pipeline(db: Session = Depends(get_db)):
    """
    Master nightly job. Runs all steps in order:
      1. Auto-close forgotten checkouts
      2. Mark absent interns
      3. Run ETL (transform.py) → refreshes features table + features.csv
      4. Retrain ML models (ml/train.py) → overwrites .pkl files

    Trigger via cron at 23:00 on weekdays:
      0 23 * * 1-5  curl -s -X POST http://localhost:8000/admin/run-nightly-pipeline

    Or call manually from the admin dashboard during development.
    """
    today   = now_local().date()
    results = _run_pipeline_for_date(today, db)
    _write_last_pipeline_run(today)

    return {
        "message": "✅ Pipeline nightly terminé",
        "ran_at":  now_local().isoformat(),
        "results": results,
    }


# ── /ml/predictions ───────────────────────────────────────────────────────────

@app.get("/ml/predictions", tags=["ML"])
def get_ml_predictions(db: Session = Depends(get_db)):
    """
    Returns the latest risk prediction and anomaly score per intern.

    - If ml/predict.py exists and models are trained → returns ML predictions.
    - Otherwise → returns rule-based labels from features.csv (cold start).

    Dashboard calls this endpoint. No ML knowledge needed on the frontend:
    just read risk_label, risk_confidence, is_anomaly, anomaly_score.

    Response shape per intern:
    {
        "intern_id":        "uuid",
        "full_name":        "Prénom Nom",
        "department":       "Cardiologie",
        "risk_label":       "Faible" | "Moyen" | "Élevé",
        "risk_confidence":  0.91,          # 0–1, null if rule-based
        "is_anomaly":       false,
        "anomaly_score":    -0.12,         # more negative = more anomalous, null if rule-based
        "source":           "ml" | "rules" # so dashboard can show a badge
    }
    """
    predict_script = os.path.join("ml", "predict.py")

    if os.path.exists(predict_script):
        # ML models exist — run predict.py and return its output
        try:
            result = subprocess.run(
                ["python", predict_script, "--output", "json"],
                capture_output=True,
                text=True,
                timeout=60,
            )
            if result.returncode == 0:
                import json
                return json.loads(result.stdout)
        except Exception:
            pass  # Fall through to rule-based

    # ── Cold start / fallback: rule-based labels from features.csv ────────
    features_path = os.path.join("etl", "features.csv")
    if not os.path.exists(features_path):
        raise HTTPException(
            status_code=503,
            detail="Pas encore de données de features. Lancez d'abord le pipeline ETL.",
        )

    import csv
    predictions = []

    # Build a quick intern lookup: id → (full_name, department_name)
    interns = db.query(Intern).all()
    departments = {d.id: d.name for d in db.query(Department).all()}
    intern_map = {
        i.id: {
            "full_name":   f"{i.first_name} {i.last_name}",
            "department":  departments.get(i.department_id, "—"),
        }
        for i in interns
    }

    with open(features_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            intern_id = row["intern_id"]
            info = intern_map.get(intern_id, {"full_name": "—", "department": "—"})
            predictions.append({
                "intern_id":       intern_id,
                "full_name":       info["full_name"],
                "department":      info["department"],
                "risk_label":      row.get("risk_label", "—"),
                "risk_confidence": None,   # not available in rule-based mode
                "is_anomaly":      False,  # not available in rule-based mode
                "anomaly_score":   None,
                "source":          "rules",
            })

    return predictions


# ── OTHER ROUTES ──────────────────────────────────────────────────────────────

@app.get("/admin/scheduler-status", tags=["Administration"])
def get_scheduler_status():
    """
    Check APScheduler status and next scheduled job times.
    Useful for verifying the nightly automation is set up correctly.
    """
    if not scheduler.running:
        return {
            "status": "❌ SCHEDULER STOPPED",
            "message": "APScheduler is not running. Restart the app.",
            "jobs": [],
        }
    
    jobs_info = []
    for job in scheduler.get_jobs():
        next_run = job.next_run_time
        next_run_local = to_local(next_run) if next_run else None
        jobs_info.append({
            "id": job.id,
            "name": job.func.__name__,
            "schedule": str(job.trigger),
            "next_run": next_run_local.isoformat() if next_run_local else "Not scheduled",
            "status": "🟢 Active",
        })
    
    return {
        "status": "🟢 SCHEDULER RUNNING",
        "message": f"{len(jobs_info)} jobs scheduled and active (Mon–Fri only)",
        "current_time": now_local().isoformat(),
        "jobs": jobs_info,
    }


@app.get("/interns")
def list_interns(
    archive_status: str | None = None,
    department_id: str | None = None,
    filiere: str | None = None,
    etablissement_id: str | None = None,
    specialite_id: str | None = None,
    year: int | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(get_db),
):
    query = db.query(Intern)
    if archive_status is not None:
        query = query.filter(Intern.archive_status == archive_status)
    if department_id is not None:
        query = query.filter(Intern.department_id == department_id)
    if filiere is not None:
        _validate_intern_filter_rules(filiere, specialite_id, year)
        query = query.filter(Intern.filiere.in_(_filiere_filter_values(filiere)))
    if etablissement_id is not None:
        query = query.filter(Intern.etablissement_id == etablissement_id)
    if specialite_id is not None:
        query = query.filter(Intern.specialite_id == specialite_id)
    if year is not None:
        query = query.filter(Intern.annee == year)

    parsed_start = _parse_optional_iso_date(start_date, "start_date")
    parsed_end = _parse_optional_iso_date(end_date, "end_date")
    _validate_rotation_dates(parsed_start, parsed_end)
    if parsed_start is not None:
        query = query.filter(Intern.end_date >= parsed_start)
    if parsed_end is not None:
        query = query.filter(Intern.start_date <= parsed_end)
    return query.all()


@app.get("/dashboard/history", response_model=list[DailyStatusOut])
def get_dashboard_data(db: Session = Depends(get_db)):
    return db.query(DailyStatus).all()


@app.post("/interns/add", tags=["Administration"])
def add_new_intern(
    intern_data: InternCreate,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    # Only DFRI and Chef de Service can add interns
    if admin.role not in (AdminRole.DFRI.value, AdminRole.CHEF_SERVICE.value):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Seuls DFRI et Chef de Service peuvent ajouter des stagiaires"
        )
    
    # Chef de Service can only add interns in their own service
    if admin.role == AdminRole.CHEF_SERVICE.value:
        if intern_data.department_id != admin.department_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Chef de Service ne peut ajouter des stagiaires que dans son service"
            )
    
    _validate_intern_business_rules(
        intern_data.filiere,
        intern_data.etablissement_id,
        intern_data.specialite_id,
        intern_data.annee,
    )

    new_uuid = str(uuid.uuid4())
    new_intern = Intern(
        id=new_uuid,
        first_name=intern_data.first_name,
        last_name=intern_data.last_name,
        department_id=intern_data.department_id,
        start_date=_parse_optional_iso_date(intern_data.start_date, "date_debut"),
        end_date=_parse_optional_iso_date(intern_data.end_date, "date_fin"),
        type=intern_data.type,
        school=intern_data.school,
        specialty=intern_data.specialty,
        filiere=intern_data.filiere,
        specialite_id=intern_data.specialite_id,
        etablissement_id=intern_data.etablissement_id,
        cne=intern_data.cne,
        annee=intern_data.annee,
        groupe=intern_data.groupe,
    )
    _validate_rotation_dates(new_intern.start_date, new_intern.end_date)
    try:
        db.add(new_intern)
        db.commit()
        db.refresh(new_intern)
        return {
            "message":   "Stagiaire ajouté avec succès",
            "id":        new_uuid,
            "full_name": f"{new_intern.first_name} {new_intern.last_name}",
        }
    except Exception as e:
        db.rollback()
        print(f"❌ REAL ERROR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── ★ EXCEL IMPORT ENDPOINTS ────────────────────────────────────────────────
def _generate_import_template():
    """Génère un fichier Excel template pour l'import de stagiaires."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stagiaires"
    
    # En-têtes
    headers = ["prenom", "nom", "departement", "type_stagiaire", "ecole", "annee_etudes", "date_debut", "date_fin", "email"]
    ws.append(headers)
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Largeurs de colonne
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 25
    
    # Retourner en mémoire
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


@app.get("/interns/import/template", tags=["Import Excel"])
def download_import_template():
    """Télécharge le template Excel pour l'import de stagiaires."""
    stream = _generate_import_template()
    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Template_Import_Stagiaires.xlsx"}
    )


@app.post("/interns/import/validate", tags=["Import Excel"])
async def validate_import_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Valide le fichier d'import sans modifier la base de données."""
    
    # Vérifier les permissions
    if admin.role not in (AdminRole.DFRI.value, AdminRole.CHEF_SERVICE.value):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Seul DFRI et Chef de Service peuvent importer des stagiaires"
        )
    
    try:
        # Lire le fichier
        contents = await file.read()
        file.file.seek(0)
        
        # Déterminer le format (xlsx ou csv)
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents), dtype=str)
        else:
            df = pd.read_excel(io.BytesIO(contents), sheet_name=0, dtype=str)
        
        # Colonnes obligatoires (en minuscules)
        df.columns = df.columns.str.lower().str.strip()
        required_cols = ['prenom', 'nom']
        
        errors_list = []
        valid_rows = []
        
        # Vérifier les colonnes obligatoires
        for col in required_cols:
            if col not in df.columns:
                errors_list.append({
                    "row": 1,
                    "message": f"Colonne obligatoire manquante : '{col}'"
                })
                break
        dept_col = 'service' if 'service' in df.columns else 'departement' if 'departement' in df.columns else None
        type_col = 'type_stagiaire' if 'type_stagiaire' in df.columns else 'type' if 'type' in df.columns else None
        school_col = 'ecole' if 'ecole' in df.columns else 'universite' if 'universite' in df.columns else None
        year_col = 'annee_etudes' if 'annee_etudes' in df.columns else 'annee' if 'annee' in df.columns else None
        email_col = 'email' if 'email' in df.columns else 'email_institutionnel' if 'email_institutionnel' in df.columns else None
        if dept_col is None:
            errors_list.append({
                "row": 1,
                "message": "Colonne obligatoire manquante : 'service'"
            })
        type_col = 'type_stagiaire' if 'type_stagiaire' in df.columns else 'type' if 'type' in df.columns else None
        school_col = 'ecole' if 'ecole' in df.columns else 'universite' if 'universite' in df.columns else None
        year_col = 'annee_etudes' if 'annee_etudes' in df.columns else 'annee' if 'annee' in df.columns else None
        email_col = 'email' if 'email' in df.columns else 'email_institutionnel' if 'email_institutionnel' in df.columns else None

        if errors_list:
            return {
                "valid_count": 0,
                "errors_count": len(errors_list),
                "errors_details": errors_list,
                "preview_rows": []
            }
        
        # Récupérer la liste des services
        departments = db.query(Department).all()
        dept_map = {d.name.lower(): d for d in departments}
        
        # Valider chaque ligne
        for idx, row in df.iterrows():
            row_num = idx + 2  # +1 pour Excel (0-indexed en python), +1 pour header
            prenom = str(row.get('prenom', '')).strip() if pd.notna(row.get('prenom')) else ''
            nom = str(row.get('nom', '')).strip() if pd.notna(row.get('nom')) else ''
            departement = str(row.get(dept_col, '')).strip() if dept_col and pd.notna(row.get(dept_col)) else ''
            intern_type = str(row.get(type_col, '')).strip() if type_col and pd.notna(row.get(type_col)) else ''
            school = str(row.get(school_col, '')).strip() if school_col and pd.notna(row.get(school_col)) else ''
            year = str(row.get(year_col, '')).strip() if year_col and pd.notna(row.get(year_col)) else ''
            email = str(row.get(email_col, '')).strip() if email_col and pd.notna(row.get(email_col)) else ''
            
            # Ignorer les lignes vides
            if not prenom and not nom and not departement:
                continue
            
            row_errors = []
            
            # Validations
            if not prenom:
                row_errors.append("Prénom vide")
            if not nom:
                row_errors.append("Nom de famille vide")
            if not departement:
                row_errors.append("Service vide")
            elif departement.lower() not in dept_map:
                row_errors.append(f"Service '{departement}' inexistant (disponibles: {', '.join([d.name for d in departments])})")
            
            if row_errors:
                errors_list.append({
                    "row": row_num,
                    "message": "; ".join(row_errors)
                })
            else:
                valid_rows.append({
                    "first_name": prenom,
                    "last_name": nom,
                    "department_name": departement,
                    "department_id": dept_map[departement.lower()].id,
                    "intern_type": intern_type,
                    "school": school,
                    "year": year,
                    "email": email,
                })
        
        # Retourner la validation
        preview = valid_rows[:5]
        return {
            "valid_count": len(valid_rows),
            "errors_count": len(errors_list),
            "errors_details": errors_list,
            "preview_rows": preview
        }
    
    except Exception as e:
        return {
            "valid_count": 0,
            "errors_count": 1,
            "errors_details": [{"row": 0, "message": f"Erreur de lecture fichier: {str(e)}"}],
            "preview_rows": []
        }


@app.post("/interns/import", tags=["Import Excel"])
async def import_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Importe les stagiaires depuis un fichier Excel/CSV."""
    
    # Vérifier les permissions
    if admin.role not in (AdminRole.DFRI.value, AdminRole.CHEF_SERVICE.value):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Seul DFRI et Chef de Service peuvent importer des stagiaires"
        )
    
    try:
        # Lire le fichier
        contents = await file.read()
        
        # Déterminer le format
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents), dtype=str)
        else:
            df = pd.read_excel(io.BytesIO(contents), sheet_name=0, dtype=str)
        
        # Normaliser les colonnes
        df.columns = df.columns.str.lower().str.strip()
        
        # Récupérer les services
        departments = db.query(Department).all()
        dept_map = {d.name.lower(): d for d in departments}
        
        dept_col = 'service' if 'service' in df.columns else 'departement' if 'departement' in df.columns else None
        type_col = 'type_stagiaire' if 'type_stagiaire' in df.columns else 'type' if 'type' in df.columns else None
        school_col = 'ecole' if 'ecole' in df.columns else 'universite' if 'universite' in df.columns else None
        year_col = 'annee_etudes' if 'annee_etudes' in df.columns else 'annee' if 'annee' in df.columns else None
        email_col = 'email' if 'email' in df.columns else 'email_institutionnel' if 'email_institutionnel' in df.columns else None
        
        created_count = 0
        failed_rows = []
        
        # Insérer les lignes valides
        for idx, row in df.iterrows():
            try:
                prenom = str(row.get('prenom', '')).strip() if pd.notna(row.get('prenom')) else ''
                nom = str(row.get('nom', '')).strip() if pd.notna(row.get('nom')) else ''
                departement = str(row.get(dept_col, '')).strip() if dept_col and pd.notna(row.get(dept_col)) else ''
                intern_type = str(row.get(type_col, '')).strip() if type_col and pd.notna(row.get(type_col)) else None
                school = str(row.get(school_col, '')).strip() if school_col and pd.notna(row.get(school_col)) else None
                year = str(row.get(year_col, '')).strip() if year_col and pd.notna(row.get(year_col)) else None
                email = str(row.get(email_col, '')).strip() if email_col and pd.notna(row.get(email_col)) else None
                
                # Ignorer les lignes vides
                if not prenom and not nom and not departement:
                    continue
                
                # Validations
                if not prenom or not nom or not departement:
                    failed_rows.append(idx + 2)
                    continue
                
                if departement.lower() not in dept_map:
                    failed_rows.append(idx + 2)
                    continue
                
                # Chef de Service peut seulement importer dans son service
                if admin.role == AdminRole.CHEF_SERVICE.value:
                    if dept_map[departement.lower()].id != admin.department_id:
                        failed_rows.append(idx + 2)
                        continue
                
                # Créer le stagiaire
                new_uuid = str(uuid.uuid4())
                new_intern = Intern(
                    id=new_uuid,
                    first_name=prenom,
                    last_name=nom,
                    department_id=dept_map[departement.lower()].id,
                    type=intern_type,
                    school=school,
                )
                db.add(new_intern)
                created_count += 1
            
            except Exception as e:
                print(f"Erreur insertion ligne {idx + 2}: {e}")
                failed_rows.append(idx + 2)
        
        # Confirmer la transaction
        db.commit()
        
        return {
            "message": f"Import terminé : {created_count} stagiaires créés",
            "created_count": created_count,
            "failed_rows": failed_rows
        }
    
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Erreur import fichier: {str(e)}"
        )


@app.get("/departments")
def list_departments(db: Session = Depends(get_db)):
    return db.query(Department).all()


@app.delete("/interns/{intern_id}", tags=["Administration"])
def delete_intern(
    intern_id: str,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    # Only DFRI and Chef de Service can delete interns
    if admin.role not in (AdminRole.DFRI.value, AdminRole.CHEF_SERVICE.value):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Seuls DFRI et Chef de Service peuvent supprimer des stagiaires"
        )
    
    intern = db.query(Intern).filter(Intern.id == intern_id).first()
    if not intern:
        raise HTTPException(status_code=404, detail="Stagiaire non trouvé")
    
    # Chef de Service can only delete interns in their own service
    if admin.role == AdminRole.CHEF_SERVICE.value:
        if intern.department_id != admin.department_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Chef de Service ne peut supprimer que les stagiaires de son service"
            )
    
    db.delete(intern)
    db.commit()
    return {"message": f"Stagiaire {intern.first_name} supprimé"}


"""class InternUpdate(BaseModel):
    first_name: str | None = None
    last_name: str | None = None
    department_id: str | None = None
    start_date: str | None = None
    end_date: str | None = None
    intern_type: str | None = None
    school: str | None = None"""


@app.patch("/interns/{intern_id}", tags=["Administration"])
def update_intern(
    intern_id: str,
    payload: InternUpdate,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin),
):
    """Update an intern's name and/or service."""
    # Only DFRI and Chef de Service can update interns
    if admin.role not in (AdminRole.DFRI.value, AdminRole.CHEF_SERVICE.value):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Seuls DFRI et Chef de Service peuvent modifier des stagiaires"
        )
    
    intern = db.query(Intern).filter(Intern.id == intern_id).first()
    if not intern:
        raise HTTPException(status_code=404, detail="Stagiaire non trouvé")
    
    # Chef de Service can only edit interns in their own service
    if admin.role == AdminRole.CHEF_SERVICE.value:
        if intern.department_id != admin.department_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Chef de Service ne peut modifier que les stagiaires de son service"
            )
        # Chef cannot change the service
        if payload.department_id is not None and payload.department_id != admin.department_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Chef de Service ne peut pas changer le service d'un stagiaire"
            )
    
    if payload.first_name is not None:
        intern.first_name = payload.first_name.strip()
    if payload.last_name is not None:
        intern.last_name = payload.last_name.strip()
    if payload.department_id is not None:
        dept = db.query(Department).filter(Department.id == payload.department_id).first()
        if not dept:
            raise HTTPException(status_code=404, detail="Service introuvable")
        intern.department_id = payload.department_id
    if payload.start_date is not None:
        intern.start_date = _parse_optional_iso_date(payload.start_date, "date_debut")
    if payload.end_date is not None:
        intern.end_date = _parse_optional_iso_date(payload.end_date, "date_fin")
    if getattr(payload, 'type', None) is not None:
        intern.type = payload.type
    if getattr(payload, 'school', None) is not None:
        intern.school = payload.school
    if getattr(payload, 'specialty', None) is not None:
        intern.specialty = payload.specialty
    if getattr(payload, 'filiere', None) is not None:
        intern.filiere = payload.filiere
    if getattr(payload, 'specialite_id', None) is not None:
        intern.specialite_id = payload.specialite_id
    if getattr(payload, 'etablissement_id', None) is not None:
        intern.etablissement_id = payload.etablissement_id
    if getattr(payload, 'cne', None) is not None:
        intern.cne = payload.cne
    if getattr(payload, 'annee', None) is not None:
        intern.annee = payload.annee
    if getattr(payload, 'groupe', None) is not None:
        intern.groupe = payload.groupe
    category_fields_touched = any(
        getattr(payload, field, None) is not None
        for field in ("filiere", "specialite_id", "etablissement_id", "annee")
    )
    if category_fields_touched:
        _validate_intern_business_rules(
            intern.filiere,
            intern.etablissement_id,
            intern.specialite_id,
            intern.annee,
        )
    _validate_rotation_dates(intern.start_date, intern.end_date)
    db.commit()
    db.refresh(intern)
    return {
        "message": f"Stagiaire mis à jour",
        "id": intern.id,
        "first_name": intern.first_name,
        "last_name": intern.last_name,
        "department_id": intern.department_id,
        "start_date": intern.start_date,
        "end_date": intern.end_date,
    }


@app.post("/interns/{intern_id}/archive", tags=["Administration"])
def archive_intern(
    intern_id: str,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin),
):
    if admin.role not in (AdminRole.DFRI.value, AdminRole.CHEF_SERVICE.value):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Seuls DFRI et Chef de Service peuvent archiver des stagiaires",
        )

    intern = db.query(Intern).filter(Intern.id == intern_id).first()
    if not intern:
        raise HTTPException(status_code=404, detail="Stagiaire non trouvé")

    if admin.role == AdminRole.CHEF_SERVICE.value and intern.department_id != admin.department_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Chef de Service ne peut archiver que les stagiaires de son service",
        )

    intern.archive_status = "archived"
    intern.archived = True
    db.commit()
    db.refresh(intern)
    return {"message": "Stagiaire archivé", "id": intern.id}


@app.patch("/interns/{intern_id}/extend", tags=["Administration"])
def extend_intern(
    intern_id: str,
    new_end_date: date_type = Body(..., embed=True),
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_super_admin),
):
    """Prolong internship end_date and reactivate if archived."""
    intern = db.query(Intern).filter(Intern.id == intern_id).first()
    if not intern:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")
    intern.end_date = str(new_end_date)
    if intern.archive_status in ("pending_archive", "archived"):
        intern.archive_status = "active"
        intern.archived = False
    db.commit()
    db.refresh(intern)
    return {"message": f"Stage prolongé jusqu'au {new_end_date}", "intern_id": intern_id}


@app.get("/rotations", tags=["Rotations"])
def list_rotations(
    department_id: Optional[str] = None,
    intern_id: Optional[str] = None,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin),
):
    """List rotations, filterable by department_id or intern_id."""
    q = db.query(Rotation)
    if department_id:
        q = q.filter(Rotation.department_id == department_id)
    if intern_id:
        q = q.filter(Rotation.intern_id == intern_id)
    rows = q.all()
    return [
        {
            "id": r.id,
            "intern_id": r.intern_id,
            "department_id": r.department_id,
            "periode_num": r.periode_num,
            "date_debut": str(r.date_debut),
            "date_fin": str(r.date_fin),
            "annee_univ": r.annee_univ,
        }
        for r in rows
    ]


@app.post("/rotations", tags=["Rotations"])
def create_rotation(
    payload: RotationCreate,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_super_admin),
):
    """Create rotation with overlap validation."""
    # Check for overlapping rotations for same intern
    overlap = db.query(Rotation).filter(
        Rotation.intern_id == payload.intern_id,
        Rotation.date_debut <= payload.date_fin,
        Rotation.date_fin >= payload.date_debut,
    ).first()
    if overlap:
        raise HTTPException(
            status_code=409,
            detail=f"Rotation chevauchée une rotation existante du {overlap.date_debut} au {overlap.date_fin}"
        )
    
    rot = Rotation(
        id=str(uuid.uuid4()),
        intern_id=payload.intern_id,
        department_id=payload.department_id,
        periode_num=payload.periode_num,
        date_debut=payload.date_debut,
        date_fin=payload.date_fin,
        annee_univ=payload.annee_univ,
    )
    db.add(rot)
    db.commit()
    db.refresh(rot)
    return {"message": "Rotation créée", "id": rot.id}


@app.get("/departments/tree", tags=["Departments"])
def get_department_tree(
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin),
):
    """Return nested department tree with parent-child relationships."""
    all_depts = db.query(Department).all()
    dept_map = {
        d.id: {
            "id": d.id,
            "name": d.name,
            "parent_id": d.parent_id,
            "children": []
        }
        for d in all_depts
    }
    roots = []
    for d in dept_map.values():
        if d["parent_id"] and d["parent_id"] in dept_map:
            dept_map[d["parent_id"]]["children"].append(d)
        else:
            roots.append(d)
    return roots


@app.patch("/departments/{dept_id}/parent", tags=["Departments"])
def set_department_parent(
    dept_id: str,
    payload: DepartmentParentUpdate,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_super_admin),
):
    """Set parent department with cyclic loop prevention."""
    dept = db.query(Department).filter(Department.id == dept_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Département introuvable")
    
    if payload.parent_id:
        # Prevent self-parent assignment
        if payload.parent_id == dept_id:
            raise HTTPException(status_code=400, detail="Un département ne peut pas être son propre parent")
        
        parent = db.query(Department).filter(Department.id == payload.parent_id).first()
        if not parent:
            raise HTTPException(status_code=404, detail="Département parent introuvable")
        
        # Prevent cyclic loops: check if parent_id is a descendant of dept_id
        current = parent
        visited = set()
        while current.parent_id:
            if current.parent_id == dept_id:
                raise HTTPException(status_code=400, detail="Assignation créerait une boucle cyclique")
            if current.parent_id in visited:
                break  # Already visited, avoid infinite loop
            visited.add(current.parent_id)
            current = db.query(Department).filter(Department.id == current.parent_id).first()
            if not current:
                break
    
    dept.parent_id = payload.parent_id
    db.commit()
    db.refresh(dept)
    return {"message": "Hiérarchie mise à jour", "id": dept.id, "parent_id": dept.parent_id}


class DepartmentUpdate(BaseModel):
    name: str


@app.patch("/departments/{dept_id}", tags=["Administration"])
def update_department(
    dept_id: str,
    payload: DepartmentUpdate,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_super_admin),
):
    """Rename a department."""
    dept = db.query(Department).filter(Department.id == dept_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Service introuvable")
    new_name = payload.name.strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Le nom ne peut pas être vide")
    conflict = db.query(Department).filter(
        Department.name == new_name, Department.id != dept_id
    ).first()
    if conflict:
        raise HTTPException(status_code=409, detail=f"Un service nommé '{new_name}' existe déjà")
    dept.name = new_name
    db.commit()
    db.refresh(dept)
    return {"message": f"Service renommé en '{new_name}'", "id": dept.id, "name": dept.name}


# ── AUTH ENDPOINTS ────────────────────────────────────────────────────────────

@app.post("/auth/login", tags=["Authentication"])
def login(request: LoginRequest, db: Session = Depends(get_db)):
    admin = db.query(Admin).filter(Admin.username == request.username).first()
    
    if not admin or not verify_password(request.password, admin.hashed_password):
        raise HTTPException(
            status_code=401,
            detail="Nom d'utilisateur ou mot de passe incorrect"
        )
    
    token = create_access_token(
        username=admin.username,
        role=AdminRole(admin.role),
        department_id=admin.department_id
    )
    
    return {
        "access_token": token,
        "token_type": "bearer",
        "role": admin.role,
        "department_id": admin.department_id
    }


@app.get("/auth/me", tags=["Authentication"])
def auth_me(current: Admin = Depends(get_current_admin)):
    return {
        "username": current.username,
        "role": current.role,
        "department_id": current.department_id,
    }


@app.post("/auth/seed-admin", tags=["Authentication"])
def seed_admin(db: Session = Depends(get_db)):
    any_admin = db.query(Admin).first()
    if any_admin:
        raise HTTPException(
            status_code=403,
            detail="Setup already completed — this endpoint is permanently locked"
        )
    
    new_admin = Admin(
        username="admin",
        hashed_password=hash_password("admin123"),
        role=AdminRole.DFRI.value,
        department_id=None
    )
    db.add(new_admin)
    db.commit()
    
    return {"message": "✅ Super admin created successfully — change your password after first login"}

@app.post("/auth/create-admin", tags=["Authentication"])
def create_admin(
    request: CreateAdminRequest,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_super_admin)   # only DFRI can create accounts
):
    """
    Create a new admin account with any of the 4 roles.
    - Only callable by a logged-in DFRI admin.
    - All roles MUST have a department_id (UUID of their department).
    - Supported roles: dfri, directeur, chef_service, secretaire
    """
    # Validate role value
    try:
        role = AdminRole(request.role)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Rôle invalide. Valeurs acceptées : {[r.value for r in AdminRole]}"
        )

    # Chef de Service must have a department_id
    if role == AdminRole.CHEF_SERVICE and not request.department_id:
        raise HTTPException(
            status_code=400,
            detail="Un Chef de Service doit être assigné à un service"
        )

    # Other roles don't use department_id
    department_id = request.department_id if role == AdminRole.CHEF_SERVICE else None

    # Check username uniqueness
    existing = db.query(Admin).filter(Admin.username == request.username).first()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"Le nom d'utilisateur '{request.username}' est déjà pris"
        )

    # Verify department exists (only for chef_service)
    if department_id:
        dept = db.query(Department).filter(Department.id == department_id).first()
        if not dept:
            raise HTTPException(
                status_code=404,
                detail="Service introuvable — vérifiez l'UUID"
            )

    new_admin = Admin(
        username=request.username,
        hashed_password=hash_password(request.password),
        role=role.value,
        department_id=department_id
    )
    db.add(new_admin)
    db.commit()

    return {
        "message":       f"✅ Compte '{request.username}' créé avec succès",
        "username":      request.username,
        "role":          role.value,
        "department_id": department_id,
    }


@app.get("/auth/admins", tags=["Authentication"])
def list_admins(
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_super_admin)
):
    """List all admin accounts (super_admin only). Passwords are never returned."""
    admins = db.query(Admin).all()
    return [
        {
            "id":            a.id,
            "username":      a.username,
            "role":          a.role,
            "department_id": a.department_id,
        }
        for a in admins
    ]


@app.delete("/auth/admins/{admin_id}", tags=["Authentication"])
def delete_admin(
    admin_id: str,
    db: Session = Depends(get_db),
    current: Admin = Depends(require_super_admin)
):
    """Delete an admin account. A super_admin cannot delete their own account."""
    if admin_id == current.id:
        raise HTTPException(
            status_code=400,
            detail="Vous ne pouvez pas supprimer votre propre compte"
        )

    target = db.query(Admin).filter(Admin.id == admin_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Compte introuvable")

    db.delete(target)
    db.commit()
    return {"message": f" Compte '{target.username}' supprimé"}


@app.post("/auth/change-password", tags=["Authentication"])
def change_password(
    request: ChangePasswordRequest,
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)   # any logged-in admin
):
    """
    Any admin can change their OWN password.
    They must provide their current password to confirm identity.
    """
    if not verify_password(request.current_password, current.hashed_password):
        raise HTTPException(
            status_code=401,
            detail="Mot de passe actuel incorrect"
        )

    if len(request.new_password) < 8:
        raise HTTPException(
            status_code=400,
            detail="Le nouveau mot de passe doit contenir au moins 8 caractères"
        )

    current.hashed_password = hash_password(request.new_password)
    db.commit()
    return {"message": " Mot de passe mis à jour avec succès"}
